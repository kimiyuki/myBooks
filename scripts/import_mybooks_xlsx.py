#!/usr/bin/env python3
"""Import the legacy MyBooks spreadsheet into a local SQLite database."""

from __future__ import annotations

import argparse
import sqlite3
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - import-time guard
    raise SystemExit(
        "openpyxl is required. Run this script with a Python environment that has openpyxl installed."
    ) from exc


EXPECTED_BOOK_HEADERS = [
    "isbn",
    "title",
    "thumbnail",
    "authors",
    "publisher",
    "pubDate",
    "url",
    "goScrap",
    "search",
    "fav",
]


@dataclass(frozen=True)
class BookRow:
    isbn: str
    title: str
    thumbnail_url: str | None
    authors: str
    publisher: str | None
    published_date: str | None
    amazon_url: str | None
    favorite: int
    registered_at: str
    updated_at: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import MyBooks.xlsx into a local SQLite database."
    )
    parser.add_argument(
        "--input",
        required=True,
        type=Path,
        help="Path to the downloaded MyBooks.xlsx file.",
    )
    parser.add_argument(
        "--output",
        default=Path("data/mybooks.db"),
        type=Path,
        help="Path to the output SQLite database.",
    )
    parser.add_argument(
        "--schema",
        default=Path("scripts/schema.sql"),
        type=Path,
        help="Path to the schema SQL file.",
    )
    parser.add_argument(
        "--replace",
        action="store_true",
        help="Replace the output database if it already exists.",
    )
    return parser.parse_args()


def normalize_isbn(value: object, row_number: int) -> str:
    if value is None:
        raise ValueError(f"row {row_number}: isbn is required")
    if isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        if not value.is_integer():
            raise ValueError(f"row {row_number}: isbn must be an integer-like value")
        text = str(int(value))
    else:
        text = str(value).strip()

    text = text.replace("-", "").strip()
    if not text:
        raise ValueError(f"row {row_number}: isbn is empty")
    return text


def normalize_required_text(value: object, field_name: str, row_number: int) -> str:
    text = normalize_optional_text(value)
    if text is None:
        raise ValueError(f"row {row_number}: {field_name} is required")
    return text


def normalize_optional_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if text == "" or text.lower() == "undefined":
        return None
    return text


def normalize_favorite(value: object, row_number: int) -> int:
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        if value in (0, 0.0):
            return 0
        if value in (1, 1.0):
            return 1
    raise ValueError(f"row {row_number}: favorite must be boolean-like")


def normalize_date_only(value: object, row_number: int, field_name: str) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text).date().isoformat()
    except ValueError as exc:
        raise ValueError(f"row {row_number}: invalid {field_name}: {text}") from exc


def normalize_datetime_text(value: object, row_number: int, field_name: str) -> str:
    if value is None:
        raise ValueError(f"row {row_number}: {field_name} is required")
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).isoformat(timespec="seconds")

    text = str(value).strip()
    if not text:
        raise ValueError(f"row {row_number}: {field_name} is empty")
    try:
        return datetime.fromisoformat(text).isoformat(timespec="seconds")
    except ValueError as exc:
        raise ValueError(f"row {row_number}: invalid {field_name}: {text}") from exc


def iter_book_rows(workbook_path: Path) -> Iterable[BookRow]:
    workbook = load_workbook(workbook_path, data_only=True)
    if "books" not in workbook.sheetnames:
        raise ValueError("books sheet is missing")

    sheet = workbook["books"]
    header_values = [sheet.cell(row=1, column=index).value for index in range(1, 11)]
    if header_values != EXPECTED_BOOK_HEADERS:
        raise ValueError(
            f"unexpected books headers: {header_values!r} != {EXPECTED_BOOK_HEADERS!r}"
        )

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if not any(cell is not None for cell in row[:11]):
            continue

        isbn = normalize_isbn(row[0], row_number)
        title = normalize_required_text(row[1], "title", row_number)
        thumbnail_url = normalize_optional_text(row[2])
        authors = normalize_required_text(row[3], "authors", row_number)
        publisher = normalize_optional_text(row[4])
        published_date = normalize_date_only(row[5], row_number, "published_date")
        favorite = normalize_favorite(row[9], row_number)
        registered_at = normalize_datetime_text(row[10], row_number, "registered_at")

        yield BookRow(
            isbn=isbn,
            title=title,
            thumbnail_url=thumbnail_url,
            authors=authors,
            publisher=publisher,
            published_date=published_date,
            amazon_url=None,
            favorite=favorite,
            registered_at=registered_at,
            updated_at=registered_at,
        )


def ensure_paths(input_path: Path, output_path: Path, schema_path: Path, replace: bool) -> None:
    if not input_path.exists():
        raise FileNotFoundError(f"input file not found: {input_path}")
    if not schema_path.exists():
        raise FileNotFoundError(f"schema file not found: {schema_path}")
    if output_path.exists() and not replace:
        raise FileExistsError(
            f"output database already exists: {output_path} (use --replace to overwrite)"
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)


def initialize_database(connection: sqlite3.Connection, schema_sql: str) -> None:
    connection.executescript(schema_sql)


def import_books(connection: sqlite3.Connection, books: Iterable[BookRow]) -> int:
    rows = [
        (
            book.isbn,
            book.title,
            book.thumbnail_url,
            book.authors,
            book.publisher,
            book.published_date,
            book.amazon_url,
            book.favorite,
            book.registered_at,
            book.updated_at,
        )
        for book in books
    ]
    connection.executemany(
        """
        INSERT INTO books (
          isbn,
          title,
          thumbnail_url,
          authors,
          publisher,
          published_date,
          amazon_url,
          favorite,
          registered_at,
          updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )
    return len(rows)


def main() -> int:
    args = parse_args()
    input_path = args.input.resolve()
    output_path = args.output.resolve()
    schema_path = args.schema.resolve()

    ensure_paths(input_path, output_path, schema_path, args.replace)

    if output_path.exists():
        output_path.unlink()

    schema_sql = schema_path.read_text(encoding="utf-8")
    books = list(iter_book_rows(input_path))

    with sqlite3.connect(output_path) as connection:
        initialize_database(connection, schema_sql)
        imported_count = import_books(connection, books)
        count = connection.execute("SELECT COUNT(*) FROM books").fetchone()[0]

    print(f"imported_books={imported_count}")
    print(f"books_count={count}")
    print(f"db_path={output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
