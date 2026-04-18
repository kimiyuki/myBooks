#!/usr/bin/env python3
"""Import legacy scrap rows and images from MyBooks.xlsx into SQLite."""

from __future__ import annotations

import argparse
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse
from urllib.request import Request, urlopen

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - import-time guard
    raise SystemExit(
        "openpyxl is required. Run this script with a Python environment that has openpyxl installed."
    ) from exc

from import_mybooks_xlsx import normalize_datetime_text, normalize_isbn


EXPECTED_SCRAP_HEADERS = [
    "isbn",
    "page",
    "picture",
    "created",
    "title",
    "page-date",
    "tweet",
]

SUPPORTED_IMAGE_TYPES = {
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
}


@dataclass(frozen=True)
class ScrapRow:
    row_number: int
    isbn: str
    page: int | None
    picture_url: str
    created_at: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import scrap rows from MyBooks.xlsx into an existing SQLite database."
    )
    parser.add_argument(
        "--input",
        required=True,
        type=Path,
        help="Path to the downloaded MyBooks.xlsx file.",
    )
    parser.add_argument(
        "--db",
        default=Path("data/mybooks.db"),
        type=Path,
        help="Path to the SQLite database.",
    )
    return parser.parse_args()


def normalize_picture_url(value: object, row_number: int) -> str:
    if value is None:
        raise ValueError(f"row {row_number}: picture is required")
    text = str(value).strip()
    if not text:
        raise ValueError(f"row {row_number}: picture is empty")
    parsed = urlparse(text)
    if parsed.scheme != "https" or not parsed.netloc:
        raise ValueError(f"row {row_number}: picture must be an https URL")
    return text


def normalize_page(value: object, row_number: int) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, str) and value.strip().lower() == "undefined":
        return None
    if isinstance(value, int):
        page = value
    elif isinstance(value, float):
        if not value.is_integer():
            raise ValueError(f"row {row_number}: page must be an integer-like value")
        page = int(value)
    else:
        page = int(str(value).strip())

    if page < 0:
        raise ValueError(f"row {row_number}: page must be 0 or greater")
    return page


def iter_scrap_rows(workbook_path: Path) -> Iterable[ScrapRow]:
    workbook = load_workbook(workbook_path, data_only=True)
    if "scraps" not in workbook.sheetnames:
        raise ValueError("scraps sheet is missing")

    sheet = workbook["scraps"]
    header_values = [sheet.cell(row=1, column=index).value for index in range(1, 8)]
    if header_values != EXPECTED_SCRAP_HEADERS:
        raise ValueError(
            f"unexpected scraps headers: {header_values!r} != {EXPECTED_SCRAP_HEADERS!r}"
        )

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if not any(cell is not None and str(cell).strip() != "" for cell in row[:7]):
            continue

        yield ScrapRow(
            row_number=row_number,
            isbn=normalize_isbn(row[0], row_number),
            page=normalize_page(row[1], row_number),
            picture_url=normalize_picture_url(row[2], row_number),
            created_at=normalize_datetime_text(row[3], row_number, "created"),
        )


def connect_db(db_path: Path) -> sqlite3.Connection:
    connection = sqlite3.connect(db_path)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    return connection


def fetch_image(url: str) -> tuple[bytes, str]:
    request = Request(
        url,
        headers={
            "User-Agent": "mybooks-import/1.0",
        },
    )
    try:
        with urlopen(request, timeout=30) as response:
            content_type = (response.headers.get_content_type() or "").lower()
            image_bytes = response.read()
    except HTTPError as exc:
        raise RuntimeError(f"failed to download image: {url} ({exc.code})") from exc
    except URLError as exc:
        raise RuntimeError(f"failed to download image: {url} ({exc.reason})") from exc

    if not image_bytes:
        raise RuntimeError(f"downloaded image is empty: {url}")
    return image_bytes, content_type


def infer_extension(image_bytes: bytes, content_type: str) -> str:
    if content_type in SUPPORTED_IMAGE_TYPES:
        return SUPPORTED_IMAGE_TYPES[content_type]
    if image_bytes.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png"
    if image_bytes.startswith(b"\xff\xd8\xff"):
        return ".jpg"
    if image_bytes.startswith(b"RIFF") and image_bytes[8:12] == b"WEBP":
        return ".webp"
    raise RuntimeError(f"unsupported image type: {content_type or 'unknown'}")


def build_relative_path(scrap: ScrapRow, extension: str) -> Path:
    created_slug = (
        scrap.created_at.replace("-", "").replace(":", "").replace("T", "T")
    )
    return Path("scraps") / scrap.isbn / f"legacy-{created_slug}-r{scrap.row_number}{extension}"


def scrap_exists(connection: sqlite3.Connection, scrap: ScrapRow) -> bool:
    return (
        connection.execute(
            """
            SELECT 1
            FROM scraps
            WHERE book_isbn = ?
              AND created_at = ?
              AND (
                (page IS NULL AND ? IS NULL)
                OR page = ?
              )
            LIMIT 1
            """,
            (scrap.isbn, scrap.created_at, scrap.page, scrap.page),
        ).fetchone()
        is not None
    )


def import_scraps(connection: sqlite3.Connection, scraps: Iterable[ScrapRow], media_root: Path) -> int:
    inserted = 0
    for scrap in scraps:
        if (
            connection.execute(
                "SELECT 1 FROM books WHERE isbn = ?",
                (scrap.isbn,),
            ).fetchone()
            is None
        ):
            raise ValueError(f"row {scrap.row_number}: book not found for isbn {scrap.isbn}")

        if scrap_exists(connection, scrap):
            continue

        image_bytes, content_type = fetch_image(scrap.picture_url)
        extension = infer_extension(image_bytes, content_type)
        relative_path = build_relative_path(scrap, extension)
        absolute_path = media_root / relative_path
        absolute_path.parent.mkdir(parents=True, exist_ok=True)
        absolute_path.write_bytes(image_bytes)

        cursor = connection.execute(
            """
            INSERT OR IGNORE INTO scraps (
              book_isbn,
              page,
              image_path,
              created_at
            ) VALUES (?, ?, ?, ?)
            """,
            (
                scrap.isbn,
                scrap.page,
                relative_path.as_posix(),
                scrap.created_at,
            ),
        )
        if cursor.rowcount == 1:
            inserted += 1

    return inserted


def main() -> int:
    args = parse_args()
    if not args.input.exists():
        raise SystemExit(f"input file not found: {args.input}")
    if not args.db.exists():
        raise SystemExit(f"database not found: {args.db}")

    scraps = list(iter_scrap_rows(args.input))
    with connect_db(args.db) as connection:
        inserted = import_scraps(connection, scraps, args.db.parent)

    print(f"Imported {inserted} scraps into {args.db}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
